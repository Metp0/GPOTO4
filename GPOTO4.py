#<3
import os
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox, Toplevel, Message

def read_gp_file(gp_file_path):
    mapping = {}
    try:
        with open(gp_file_path, "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    key, values = line.strip().split("=")
                    key = key.strip().lstrip("-")
                    codes = [v.strip().strip(";").upper() for v in values.split(",")]
                    for code in codes:
                        mapping[code] = key
    except FileNotFoundError:
        messagebox.showerror("Hata", f"'{gp_file_path}' dosyası bulunamadı.")
    return mapping

def process_excel(file_path):
    gp_path = os.path.join(os.getcwd(), "GT.txt")
    code_map = read_gp_file(gp_path)
    if not code_map:
        return

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.insert_cols(7)

        header_cell = ws["G1"]
        header_cell.value = "Bölüm Adı"
        fill_color = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # %60 vurgu 6 yeşil tonu
        header_cell.fill = fill_color

        for row in range(2, ws.max_row + 1):
            cell_value = ws[f"F{row}"].value
            if cell_value:
                code = str(cell_value).strip().upper()
                category = code_map.get(code, "")
                ws[f"G{row}"] = category

        wb.save(file_path)
        messagebox.showinfo("Başarılı", "Excel başarıyla işlendi!")
    except Exception as e:
        messagebox.showerror("Hata", f"İşlem sırasında bir hata oluştu:\n{e}")

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
    if file_path:
        file_entry.delete(0, 'end')
        file_entry.insert(0, file_path)

def clear_path():
    file_entry.delete(0, 'end')

def open_help_window():
    help_win = Toplevel(app)
    help_win.title("Yardım")
    help_win.geometry("400x400")
    help_win.resizable(False, False)
    help_text = (
        "Program, seçilen Excel dosyasında 'F' sütunundaki GTxxx verisine göre, karşılık gelen veriyi 'G' sütununa ekler. Bu veriler, programla aynı dizinde bulunan 'GT.txt' dosyasından alınır. 'GT.txt' dosyası, mevcut formata uygun olduğu sürece düzenlenebilir; dosyaya veri eklenip çıkarılabilir."
    )

    message_widget = Message(help_win, text=help_text, font=("Segoe UI", 12), width=360)
    message_widget.pack(expand=True, padx=20, pady=20)

app = tb.Window(themename="flatly")
app.title("GT'ye Göre Bölüm Adı")
app.geometry("950x250")
app.resizable(False, False)

font_style = ("Segoe UI", 11, "bold")
style = tb.Style()

style.configure("success.TButton", font=font_style)
style.configure("danger.TButton", font=font_style)
style.configure("warning.TButton", font=font_style)

frame = tb.Frame(app, padding=20)
frame.pack(expand=True)

label = tb.Label(frame, text="Headcount Exceli:", font=font_style)
label.grid(row=0, column=0, padx=10, pady=10, sticky="e")

file_entry = tb.Entry(frame, width=55, font=font_style)
file_entry.grid(row=0, column=1, padx=10, pady=10, sticky="w")

select_button = tb.Button(
    frame, text="Seç", bootstyle="warning", width=10, command=select_file,
)
select_button.grid(row=0, column=2, padx=10, pady=10)

process_button = tb.Button(
    frame, text="İşlemi Başlat", bootstyle="success", width=20,
    command=lambda: process_excel(file_entry.get())
)
process_button.grid(row=1, column=1, padx=10, pady=60, sticky="e")

clear_button = tb.Button(
    frame, text="Temizle", bootstyle="danger", width=20,
    command=clear_path
)
clear_button.grid(row=1, column=1, padx=10, pady=60, sticky="w")

help_button = tb.Button(
    app, text="?", bootstyle="info", width=3, command=open_help_window
)
help_button.place(x=10, y=215)


app.mainloop()